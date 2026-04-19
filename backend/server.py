from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from fastapi import FastAPI, APIRouter, HTTPException, Request, Depends, UploadFile, File
from fastapi.responses import Response, StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.gzip import GZipMiddleware
from starlette.middleware.cors import CORSMiddleware
import asyncpg, aiofiles, uuid, asyncio, io
import os, logging, bcrypt, jwt, secrets, string, json, math
import ssl
import certifi
import httpx
from urllib.parse import urlparse
from datetime import datetime, timezone, timedelta
from pydantic import BaseModel
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from modules.common.indexes import create_indexes
from modules.common.logging import configure_logging
from modules.common.middleware import install_middleware

logger = configure_logging()

ORDER_BILLING_STEP_SQM = 0.5
KEEP_ALIVE_INTERVAL_SECONDS = 60
INVENTORY_SYNC_STATUSES = {"tayyor", "yetkazilmoqda", "yetkazildi"}


def load_database_url() -> str:
    """Render/Supabase uchun DSN: bo'sh hostname va IDNA xatolarini oldini oladi."""
    raw = os.environ.get("DATABASE_URL")
    if raw is None or not str(raw).strip():
        raise RuntimeError(
            "DATABASE_URL yo'q yoki bo'sh. Render → Environment → DATABASE_URL qo'shing "
            "(Supabase → Database → URI, Transaction pooler, port 6543)."
        )
    url = str(raw).strip()
    if len(url) >= 2 and url[0] == url[-1] and url[0] in "\"'":
        url = url[1:-1].strip()
    url = url.lstrip("\ufeff\u200b\u200c\u200d").strip()
    parsed = urlparse(url.replace("postgres://", "postgresql://", 1))
    host = parsed.hostname
    if not host or not str(host).strip():
        raise RuntimeError(
            "DATABASE_URL da HOSTNAME bo'sh (noto'g'ri URL). Odatda parolda @ : # % bo'lganda "
            "URL-encoding qilinmagan: urllib.parse.quote_plus(parol) bilan almashtiring. "
            "Yoki Supabase'dan 'Connection string' ni to'liq nusxalang."
        )
    hn = str(host).strip().lower()
    # Railway/Render da ko'pincha hujjatdan "..." yoki [YOUR-PASSWORD] nusxalanadi
    if hn == "..." or hn.startswith("...") or "..." in hn:
        raise RuntimeError(
            "DATABASE_URL ichida hostname o'rniga '...' qolgan (namuna matn). "
            "Railway → Variables → DATABASE_URL: Supabase → Project Settings → Database → "
            "Connection string → URI (Transaction pooler, port 6543) ni BUTUNLAY nusxalang. "
            "Hech qayerda ... yoki [YOUR-PASSWORD] qoldirmang."
        )
    if hn.startswith(".") or ".." in hn or hn.startswith("@"):
        raise RuntimeError(
            f"DATABASE_URL hostname noto'g'ri: {host!r}. "
            "URL noto'g'ri kesilgan yoki boshida nuqta bor — Supabase URI ni qayta nusxalang."
        )
    return url


def asyncpg_ssl_context_for_dsn(dsn: str) -> Optional[ssl.SSLContext]:
    """
    Supabase pooler: Railway/Render yo'lida CA zanjiri ba'zan 'self-signed in chain' beradi.
    Standart: TLS shifrlangan, lekin server sertifikati tekshirilmaydi (asyncpg + bulut uchun odatiy yechim).
    Qat'iy tekshiruv: ASYNCPG_STRICT_SSL=1 (mahalliy/yaxshi tarmoq).
    """
    try:
        h = (urlparse(dsn.replace("postgres://", "postgresql://", 1)).hostname or "").lower()
    except Exception:
        return None
    if "supabase.co" not in h:
        return None
    strict = os.environ.get("ASYNCPG_STRICT_SSL", "").strip().lower() in ("1", "true", "yes")
    if strict:
        return ssl.create_default_context(cafile=certifi.where())
    ctx = ssl.SSLContext(ssl.PROTOCOL_TLS_CLIENT)
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    logger.debug("Supabase pooler: TLS (sertifikat verify o'chiq, bulut muvofiqligi)")
    return ctx


def asyncpg_pool_kwargs():
    kw = dict(
        min_size=1,
        max_size=10,
        statement_cache_size=0,
        command_timeout=60,
        timeout=20,
        max_inactive_connection_lifetime=120,
    )
    ctx = asyncpg_ssl_context_for_dsn(DATABASE_URL)
    if ctx is not None:
        kw["ssl"] = ctx
    return kw


DATABASE_URL = load_database_url()

UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

app = FastAPI()
app.mount("/api/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")
api_router = APIRouter(prefix="/api")
JWT_ALGORITHM = "HS256"
install_middleware(app, logger)

# ─── IN-MEMORY CACHE (DB tezlashtirish) ───
import time as _time

class MemoryCache:
    """Oddiy xotiradagi kesh — DB so'rovlarni 3-5x tezlashtiradi"""
    def __init__(self):
        self._store: dict = {}
        self._ttl: dict = {}

    def get(self, key: str):
        if key in self._store and _time.time() < self._ttl.get(key, 0):
            return self._store[key]
        self._store.pop(key, None)
        self._ttl.pop(key, None)
        return None

    def set(self, key: str, value, ttl_seconds: int = 30):
        self._store[key] = value
        self._ttl[key] = _time.time() + ttl_seconds

    def invalidate(self, *prefixes):
        """Berilgan prefikslar bilan boshlanadigan barcha keshlarni o'chiradi"""
        keys_to_del = [k for k in self._store if any(k.startswith(p) for p in prefixes)]
        for k in keys_to_del:
            self._store.pop(k, None)
            self._ttl.pop(k, None)

    def clear(self):
        self._store.clear()
        self._ttl.clear()

cache = MemoryCache()
schema_cache = {
    "tables": {},
    "columns": {},
    "types": {},
    "enums": {},
}

# ─── DB Pool ───
pool: asyncpg.Pool = None

async def get_pool() -> asyncpg.Pool:
    global pool
    if pool is None or pool._closed:
        pool = await asyncpg.create_pool(DATABASE_URL, **asyncpg_pool_kwargs())
    return pool

# ─── Helpers ───
def get_jwt_secret(): return os.environ["JWT_SECRET"]
def hash_password(pw: str) -> str: return bcrypt.hashpw(pw.encode(), bcrypt.gensalt(4)).decode()
def verify_password(plain: str, hashed: str) -> bool: return bcrypt.checkpw(plain.encode(), hashed.encode())

def create_access_token(uid: str, email: str, role: str) -> str:
    return jwt.encode({"sub": uid, "email": email, "role": role, "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "access"}, get_jwt_secret(), algorithm=JWT_ALGORITHM)

def generate_order_code():
    return ''.join(secrets.choice(string.ascii_uppercase + string.digits) for _ in range(8))

def row_to_dict(row):
    if row is None:
        return None
    return dict(row)


def round_money(value: float) -> float:
    return round(float(value or 0) + 1e-9, 2)


def to_billable_sqm(area: float) -> float:
    area = max(float(area or 0), 0.0)
    if area <= 0:
        return 0.0
    return round(math.ceil((area - 1e-9) / ORDER_BILLING_STEP_SQM) * ORDER_BILLING_STEP_SQM, 2)


def parse_json_field(value, default=None):
    if value in (None, ""):
        return default
    if isinstance(value, str):
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            return default if default is not None else value
    return value


async def table_exists(db, table_name: str) -> bool:
    cached = schema_cache["tables"].get(table_name)
    if cached is not None:
        return cached
    exists = bool(await db.fetchval(
        """
        SELECT EXISTS (
            SELECT 1
            FROM information_schema.tables
            WHERE table_schema = 'public' AND table_name = $1
        )
        """,
        table_name,
    ))
    schema_cache["tables"][table_name] = exists
    return exists


async def column_exists(db, table_name: str, column_name: str) -> bool:
    cache_key = f"{table_name}.{column_name}"
    cached = schema_cache["columns"].get(cache_key)
    if cached is not None:
        return cached
    exists = bool(await db.fetchval(
        """
        SELECT EXISTS (
            SELECT 1
            FROM information_schema.columns
            WHERE table_schema = 'public' AND table_name = $1 AND column_name = $2
        )
        """,
        table_name,
        column_name,
    ))
    schema_cache["columns"][cache_key] = exists
    return exists


async def column_udt_name(db, table_name: str, column_name: str) -> Optional[str]:
    cache_key = f"{table_name}.{column_name}"
    if cache_key in schema_cache["types"]:
        return schema_cache["types"][cache_key]
    udt_name = await db.fetchval(
        """
        SELECT udt_name
        FROM information_schema.columns
        WHERE table_schema = 'public' AND table_name = $1 AND column_name = $2
        LIMIT 1
        """,
        table_name,
        column_name,
    )
    schema_cache["types"][cache_key] = udt_name
    return udt_name


async def enum_values_for_column(db, table_name: str, column_name: str) -> List[str]:
    cache_key = f"{table_name}.{column_name}"
    cached = schema_cache["enums"].get(cache_key)
    if cached is not None:
        return cached
    rows = await db.fetch(
        """
        SELECT e.enumlabel
        FROM information_schema.columns c
        JOIN pg_type t ON t.typname = c.udt_name
        JOIN pg_enum e ON e.enumtypid = t.oid
        WHERE c.table_schema = 'public' AND c.table_name = $1 AND c.column_name = $2
        ORDER BY e.enumsortorder
        """,
        table_name,
        column_name,
    )
    values = [str(row["enumlabel"]) for row in rows]
    schema_cache["enums"][cache_key] = values
    return values


def build_insert_statement(table_name: str, values: dict) -> tuple[str, list]:
    columns = list(values.keys())
    placeholders = [f"${idx}" for idx in range(1, len(columns) + 1)]
    sql = f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES ({', '.join(placeholders)}) RETURNING *"
    return sql, [values[column] for column in columns]


async def filter_existing_fields(db, table_name: str, values: dict) -> dict:
    result = {}
    for field, value in values.items():
        if await column_exists(db, table_name, field):
            result[field] = value
    return result


async def resolve_profile_id_for_user(db, user: dict) -> Optional[str]:
    email = str(user.get("email") or "").strip().lower()
    if not email or not await table_exists(db, "profiles"):
        return None
    profile_id = await db.fetchval(
        "SELECT id::text FROM profiles WHERE lower(email) = $1 LIMIT 1",
        email,
    )
    return str(profile_id) if profile_id else None


async def resolve_actor_ids(db, user: dict) -> List[str]:
    ids: List[str] = []
    user_id = str(user.get("id") or "").strip()
    if user_id:
        ids.append(user_id)
    profile_id = await resolve_profile_id_for_user(db, user)
    if profile_id and profile_id not in ids:
        ids.append(profile_id)
    return ids


def pick_reference_value(id_candidates: List[str], udt_name: Optional[str]):
    candidates = [str(value).strip() for value in id_candidates if str(value).strip()]
    if not candidates:
        return None
    if udt_name == "uuid":
        for value in reversed(candidates):
            if "-" in value:
                return value
        return None
    if udt_name in {"int2", "int4", "int8"}:
        for value in candidates:
            if value.lstrip("-").isdigit():
                return int(value)
        return None
    return candidates[0]


async def load_user_name_map(db, id_texts: List[str]) -> dict:
    ids = sorted({str(value).strip() for value in id_texts if str(value).strip()})
    if not ids:
        return {}
    names: dict = {}
    if await table_exists(db, "users"):
        rows = await db.fetch(
            """
            SELECT id::text AS lookup_id, COALESCE(NULLIF(name, ''), email, 'User') AS display_name
            FROM users
            WHERE id::text = ANY($1::text[])
            """,
            ids,
        )
        for row in rows:
            names[str(row["lookup_id"])] = row["display_name"]
    if await table_exists(db, "profiles"):
        rows = await db.fetch(
            """
            SELECT id::text AS lookup_id, COALESCE(NULLIF(name, ''), email, 'User') AS display_name
            FROM profiles
            WHERE id::text = ANY($1::text[])
            """,
            ids,
        )
        for row in rows:
            names[str(row["lookup_id"])] = row["display_name"]
    return names


async def get_people_table(db) -> Optional[str]:
    if await table_exists(db, "users"):
        return "users"
    if await table_exists(db, "profiles"):
        return "profiles"
    return None


async def normalize_status_for_db(db, desired_status: str) -> str:
    allowed = await enum_values_for_column(db, "orders", "status")
    if not allowed or desired_status in allowed:
        return desired_status
    fallbacks = {
        "kutilmoqda": ["kutilmoqda"],
        "tasdiqlangan": ["tasdiqlangan", "tayyorlanmoqda", "kutilmoqda"],
        "tayyorlanmoqda": ["tayyorlanmoqda", "tasdiqlangan", "kutilmoqda"],
        "tayyor": ["tayyor", "yetkazildi"],
        "yetkazilmoqda": ["yetkazilmoqda", "tayyor", "yetkazildi"],
        "yetkazildi": ["yetkazildi", "tayyor"],
        "rad_etilgan": ["rad_etilgan", "kutilmoqda"],
    }
    for candidate in fallbacks.get(desired_status, [desired_status]):
        if candidate in allowed:
            return candidate
    return allowed[0]


async def fetch_order_row(db, order_id: str):
    return await db.fetchrow("SELECT * FROM orders WHERE id::text = $1", str(order_id))


async def update_order_fields(conn, order_id: str, values: dict) -> None:
    updates = []
    params = []
    idx = 1
    for field, value in values.items():
        if await column_exists(conn, "orders", field):
            updates.append(f"{field} = ${idx}")
            params.append(value)
            idx += 1
    if not updates:
        return
    params.append(str(order_id))
    await conn.execute(
        f"UPDATE orders SET {', '.join(updates)} WHERE id::text = ${idx}",
        *params,
    )


async def fetch_order_items_map(db, order_rows: List[dict]) -> dict:
    orders = [row_to_dict(row) for row in order_rows if row is not None]
    order_ids = [str(order["id"]) for order in orders if order.get("id") is not None]
    if not order_ids:
        return {}

    if await column_exists(db, "orders", "items"):
        result = {}
        for order in orders:
            raw_items = parse_json_field(order.get("items"), default=[]) or []
            normalized_items = []
            for item in raw_items:
                current = dict(item)
                if current.get("material_id") is not None:
                    current["material_id"] = str(current["material_id"])
                if current.get("assigned_worker_id") not in (None, ""):
                    current["assigned_worker_id"] = str(current["assigned_worker_id"])
                normalized_items.append(current)
            result[str(order["id"])] = normalized_items
        return result

    if not await table_exists(db, "order_items"):
        return {order_id: [] for order_id in order_ids}

    worker_refs_by_order = {
        str(order["id"]): str(order["worker_id"])
        for order in orders
        if order.get("worker_id") not in (None, "")
    }
    worker_names = await load_user_name_map(db, list(worker_refs_by_order.values()))
    rows = await db.fetch(
        """
        SELECT *
        FROM order_items
        WHERE order_id::text = ANY($1::text[])
        ORDER BY order_id::text ASC, COALESCE(item_index, 0) ASC, created_at ASC
        """,
        order_ids,
    )

    result = {order_id: [] for order_id in order_ids}
    for row in rows:
        item = row_to_dict(row)
        order_id = str(item["order_id"])
        assigned_worker_id = worker_refs_by_order.get(order_id, "")
        result.setdefault(order_id, []).append({
            "id": str(item["id"]),
            "material_id": str(item["material_id"]) if item.get("material_id") is not None else "",
            "material_name": item.get("material_name", ""),
            "width": float(item.get("width") or 0),
            "height": float(item.get("height") or 0),
            "quantity": int(item.get("quantity") or 1),
            "exact_sqm": round(float(item.get("sqm") or 0), 4),
            "billable_sqm": round(float(item.get("sqm") or 0), 2),
            "sqm": round(float(item.get("sqm") or 0), 2),
            "price_per_sqm": float(item.get("unit_price") or 0),
            "price": round_money(item.get("total_price") or 0),
            "total_price": round_money(item.get("total_price") or 0),
            "notes": item.get("notes") or "",
            "assigned_worker_id": assigned_worker_id,
            "assigned_worker_name": worker_names.get(assigned_worker_id, ""),
            "worker_status": item.get("worker_status") or ("assigned" if assigned_worker_id else "pending"),
        })
    return result


async def serialize_orders(db, rows) -> List[dict]:
    order_rows = [row_to_dict(row) for row in rows if row is not None]
    if not order_rows:
        return []

    items_map = await fetch_order_items_map(db, order_rows)
    lookup_ids = []
    for order in order_rows:
        if order.get("dealer_id") not in (None, ""):
            lookup_ids.append(str(order["dealer_id"]))
        if order.get("worker_id") not in (None, ""):
            lookup_ids.append(str(order["worker_id"]))
    for items in items_map.values():
        for item in items:
            if item.get("assigned_worker_id"):
                lookup_ids.append(str(item["assigned_worker_id"]))
    user_names = await load_user_name_map(db, lookup_ids)

    out = []
    for order in order_rows:
        order_id = str(order["id"])
        dealer_id = str(order["dealer_id"]) if order.get("dealer_id") is not None else ""
        worker_id = str(order["worker_id"]) if order.get("worker_id") is not None else ""
        delivery_info = parse_json_field(order.get("delivery_info"), default=None)
        items = items_map.get(order_id, [])

        for item in items:
            fallback_worker_id = item.get("assigned_worker_id") or worker_id
            if fallback_worker_id:
                item["assigned_worker_id"] = str(fallback_worker_id)
                item["assigned_worker_name"] = item.get("assigned_worker_name") or user_names.get(str(fallback_worker_id), "")
                if item.get("worker_status") in (None, "", "pending"):
                    item["worker_status"] = "assigned"

        current = dict(order)
        current["id"] = order_id
        current["dealer_id"] = dealer_id
        current["dealer_name"] = current.get("dealer_name") or user_names.get(dealer_id, "")
        current["items"] = items
        current["delivery_info"] = delivery_info
        if worker_id:
            current["worker_id"] = worker_id
        out.append(current)

    return out


async def serialize_order(db, row):
    orders = await serialize_orders(db, [row] if row is not None else [])
    return orders[0] if orders else None


def serialize_order_record(row):
    order = row_to_dict(row)
    if order is None:
        return None
    order["id"] = str(order["id"])
    if order.get("dealer_id") is not None:
        order["dealer_id"] = str(order["dealer_id"])
    if order.get("worker_id") is not None:
        order["worker_id"] = str(order["worker_id"])
    order["items"] = parse_json_field(order.get("items"), default=[]) or []
    order["delivery_info"] = parse_json_field(order.get("delivery_info"), default=None)
    return order


def format_export_datetime(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M")
    text = str(value)
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        return parsed.strftime("%Y-%m-%d %H:%M")
    except ValueError:
        return text[:16].replace("T", " ")


def build_order_item(it) -> dict:
    exact_sqm = max(float(it.width), 0) * max(float(it.height), 0) * max(int(it.quantity), 0)
    billable_sqm = to_billable_sqm(exact_sqm)
    price = round_money(billable_sqm * float(it.price_per_sqm))
    return {
        "material_id": str(it.material_id),
        "material_name": it.material_name,
        "width": float(it.width),
        "height": float(it.height),
        "quantity": int(it.quantity),
        "exact_sqm": round(exact_sqm, 4),
        "billable_sqm": billable_sqm,
        "sqm": billable_sqm,
        "price_per_sqm": float(it.price_per_sqm),
        "price": price,
        "notes": getattr(it, "notes", ""),
        "assigned_worker_id": "",
        "assigned_worker_name": "",
        "worker_status": "pending",
    }


async def deduct_inventory_for_order(conn, order_row, now: str):
    order = row_to_dict(order_row)
    if order is None:
        raise HTTPException(404, "Buyurtma topilmadi")
    if order.get("inventory_deducted"):
        return

    order_id = str(order["id"])
    if not await column_exists(conn, "materials", "stock_quantity"):
        await update_order_fields(conn, order_id, {"inventory_deducted": True, "updated_at": now})
        return

    if await column_exists(conn, "orders", "items"):
        items = parse_json_field(order.get("items"), default=[]) or []
    elif await table_exists(conn, "order_items"):
        item_rows = await conn.fetch(
            "SELECT material_id, sqm FROM order_items WHERE order_id::text = $1",
            order_id,
        )
        items = [
            {
                "material_id": row["material_id"],
                "billable_sqm": row["sqm"],
                "sqm": row["sqm"],
            }
            for row in item_rows
        ]
    else:
        items = []

    required_by_material: dict[str, float] = {}

    for item in items:
        material_id = item.get("material_id")
        sqm = float(item.get("billable_sqm") or item.get("sqm") or 0)
        if not material_id or sqm <= 0:
            continue
        mid = str(material_id)
        required_by_material[mid] = round_money(required_by_material.get(mid, 0) + sqm)

    if not required_by_material:
        await update_order_fields(conn, order_id, {"inventory_deducted": True, "updated_at": now})
        return

    material_ids = list(required_by_material.keys())
    material_rows = await conn.fetch(
        """
        SELECT id::text AS lookup_id, name, stock_quantity, COALESCE(unit, 'kv.m') AS unit
        FROM materials
        WHERE id::text = ANY($1::text[])
        """,
        material_ids,
    )
    materials = {str(row["lookup_id"]): row for row in material_rows}
    shortages = []

    for material_id, required_sqm in required_by_material.items():
        material = materials.get(material_id)
        if material is None:
            shortages.append(f"Material topilmadi: #{material_id}")
            continue
        available = float(material["stock_quantity"] or 0)
        if available + 1e-9 < required_sqm:
            shortages.append(
                f"{material['name']} omborda yetarli emas ({round_money(available)} {material['unit']} bor, {required_sqm} kerak)"
            )

    if shortages:
        raise HTTPException(400, "; ".join(shortages))

    for material_id, required_sqm in required_by_material.items():
        await conn.execute(
            "UPDATE materials SET stock_quantity = stock_quantity - $1 WHERE id::text = $2",
            required_sqm,
            material_id,
        )

    await update_order_fields(conn, order_id, {"inventory_deducted": True, "updated_at": now})

async def get_current_user(request: Request) -> dict:
    auth = request.headers.get("Authorization", "")
    if not auth.startswith("Bearer "): 
        logger.error("Auth error: Missing or invalid Authorization header")
        raise HTTPException(401, "Not authenticated")
    try:
        p = jwt.decode(auth[7:], get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        try:
            user_id = int(p["sub"])
        except (ValueError, TypeError):
            raise HTTPException(401, "Invalid token format - please login again")
            
        cache_key = f"user_auth_{user_id}"
        cached_user = cache.get(cache_key)
        if cached_user:
            return cached_user
            
        db = await get_pool()
        row = await db.fetchrow("SELECT * FROM users WHERE id = $1", user_id)
        if not row: raise HTTPException(401, "User not found")
        user = row_to_dict(row)
        user["id"] = str(user["id"])
        user.pop("password_hash", None)
        
        cache.set(cache_key, user, 60) # 1 minut xotirada saqlash
        return user
    except jwt.ExpiredSignatureError: 
        logger.error("Auth error: Token expired")
        raise HTTPException(401, "Token expired")
    except jwt.InvalidTokenError as e: 
        logger.error(f"Auth error: Invalid token {e}")
        raise HTTPException(401, "Invalid token")
    except HTTPException: raise
    except Exception as e:
        logger.error(f"Auth error (other): {e}")
        raise HTTPException(401, "Authentication failed")

async def require_admin(request: Request) -> dict:
    u = await get_current_user(request)
    if u.get("role") != "admin": raise HTTPException(403, "Admin only")
    return u

async def require_worker(request: Request) -> dict:
    u = await get_current_user(request)
    if u.get("role") != "worker": raise HTTPException(403, "Worker only")
    return u

# ─── Pydantic Models ───
class LoginReq(BaseModel): email: str; password: str
class DealerCreate(BaseModel): name: str; email: str; password: str; phone: str = ""; address: str = ""; credit_limit: float = 0
class DealerUpdate(BaseModel): name: Optional[str] = None; phone: Optional[str] = None; address: Optional[str] = None; credit_limit: Optional[float] = None
class WorkerCreate(BaseModel): name: str; email: str; password: str; phone: str = ""; specialty: str = ""
class MaterialCreate(BaseModel): name: str; category: str = ""; category_id: Optional[int] = None; price_per_sqm: float; stock_quantity: float; unit: str = "kv.m"; description: str = ""; image_url: str = ""
class MaterialUpdate(BaseModel): name: Optional[str] = None; category: Optional[str] = None; category_id: Optional[int] = None; price_per_sqm: Optional[float] = None; stock_quantity: Optional[float] = None; description: Optional[str] = None; image_url: Optional[str] = None
class CategoryCreate(BaseModel): name: str; description: str = ""; image_url: str = ""
class CategoryUpdate(BaseModel): name: Optional[str] = None; description: Optional[str] = None; image_url: Optional[str] = None
class OrderItemCreate(BaseModel): material_id: str; material_name: str; width: float; height: float; quantity: int = 1; price_per_sqm: float; notes: str = ""
class OrderCreate(BaseModel): items: List[OrderItemCreate]; notes: str = ""
class OrderStatusUpdate(BaseModel): status: str; rejection_reason: str = ""
class MessageCreate(BaseModel): receiver_id: str; text: str
class AssignItemReq(BaseModel): worker_id: str
class DeliveryInfoReq(BaseModel): driver_name: str; driver_phone: str; plate_number: str = ""
class PaymentCreate(BaseModel): amount: float; note: str = ""

# ─── EXCHANGE RATE (Real-time USD/UZS) ───
@api_router.get("/exchange-rate")
async def get_exchange_rate():
    """O'zbekiston Markaziy Banki dan real vaqtda dollar kursini olish"""
    cached = cache.get("exchange_rate")
    if cached: return cached
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.get("https://cbu.uz/oz/arkhiv-kursov-valyut/json/USD/")
            data = resp.json()
            if data and len(data) > 0:
                rate = float(data[0]["Rate"])
                result = {"rate": rate, "currency": "UZS", "date": data[0].get("Date", ""), "source": "CBU.uz"}
                cache.set("exchange_rate", result, 3600)
                return result
    except Exception as e:
        logger.warning(f"CBU kurs olishda xatolik: {e}")
    fallback = {"rate": 12800.0, "currency": "UZS", "date": "", "source": "fallback"}
    cache.set("exchange_rate", fallback, 300)
    return fallback

# ─── AUTH ───
@api_router.post("/auth/login")
async def login(req: LoginReq):
    db = await get_pool()
    user = await db.fetchrow("SELECT * FROM users WHERE email = $1", req.email.strip().lower())
    
    if not user:
        raise HTTPException(401, "Email yoki parol noto'g'ri")
        
    loop = asyncio.get_running_loop()
    is_valid = await loop.run_in_executor(None, verify_password, req.password, user["password_hash"])
    
    if not is_valid:
        raise HTTPException(401, "Email yoki parol noto'g'ri")
    token = create_access_token(str(user["id"]), user["email"], user["role"])
    u = row_to_dict(user)
    u["id"] = str(u["id"])
    u.pop("password_hash", None)
    return {"token": token, "user": u}

@api_router.get("/auth/me")
async def get_me(user: dict = Depends(get_current_user)): return {"user": user}

# ─── AUTH: Update own profile ───
@api_router.put("/auth/profile")
async def update_profile(request: Request, user: dict = Depends(get_current_user)):
    body = await request.json()
    new_email = body.get("email", "").strip().lower()
    new_password = body.get("password", "").strip()
    current_password = body.get("current_password", "").strip()
    if not current_password:
        raise HTTPException(400, "Joriy parolni kiriting")
    db = await get_pool()
    db_user = await db.fetchrow("SELECT * FROM users WHERE id = $1", int(user["id"]))
    if not db_user or not verify_password(current_password, db_user["password_hash"]):
        raise HTTPException(400, "Joriy parol noto'g'ri")
    updates = []
    params = []
    param_idx = 1
    if new_email and new_email != db_user["email"]:
        existing = await db.fetchrow("SELECT id FROM users WHERE email = $1 AND id != $2", new_email, int(user["id"]))
        if existing:
            raise HTTPException(400, "Bu email allaqachon mavjud")
        updates.append(f"email = ${param_idx}")
        params.append(new_email)
        param_idx += 1
    if new_password:
        if len(new_password) < 4:
            raise HTTPException(400, "Parol kamida 4 ta belgi")
        updates.append(f"password_hash = ${param_idx}")
        params.append(hash_password(new_password))
        param_idx += 1
    if not updates:
        raise HTTPException(400, "O'zgartirish yo'q")
    params.append(int(user["id"]))
    await db.execute(f"UPDATE users SET {', '.join(updates)} WHERE id = ${param_idx}", *params)
    updated = await db.fetchrow("SELECT * FROM users WHERE id = $1", int(user["id"]))
    u = row_to_dict(updated)
    u["id"] = str(u["id"])
    u.pop("password_hash", None)
    token = create_access_token(u["id"], u.get("email", ""), u.get("role", ""))
    return {"user": u, "token": token, "message": "Profil yangilandi"}

# ─── DEALERS ───
@api_router.post("/dealers")
async def create_dealer(d: DealerCreate, admin: dict = Depends(require_admin)):
    db = await get_pool()
    existing = await db.fetchrow("SELECT id FROM users WHERE email = $1", d.email.strip().lower())
    if existing: raise HTTPException(400, "Email mavjud")
    now = datetime.now(timezone.utc).isoformat()
    row = await db.fetchrow(
        "INSERT INTO users (name, email, password_hash, role, phone, address, credit_limit, debt, specialty, created_at) VALUES ($1,$2,$3,$4,$5,$6,$7,0,'',$8) RETURNING *",
        d.name, d.email.strip().lower(), hash_password(d.password), "dealer", d.phone, d.address, d.credit_limit, now
    )
    u = row_to_dict(row)
    u["id"] = str(u["id"])
    u.pop("password_hash", None)
    cache.invalidate("dealers")
    return u

@api_router.get("/dealers")
async def list_dealers(admin: dict = Depends(require_admin)):
    cached = cache.get("dealers_list")
    if cached: return cached
    db = await get_pool()
    rows = await db.fetch("SELECT * FROM users WHERE role = 'dealer' ORDER BY created_at DESC")
    out = []
    for r in rows:
        u = row_to_dict(r)
        u["id"] = str(u["id"])
        u.pop("password_hash", None)
        out.append(u)
    cache.set("dealers_list", out, 30)
    return out

@api_router.put("/dealers/{did}")
async def update_dealer(did: str, data: DealerUpdate, admin: dict = Depends(require_admin)):
    db = await get_pool()
    updates = []
    params = []
    idx = 1
    for field in ["name", "phone", "address", "credit_limit"]:
        val = getattr(data, field, None)
        if val is not None:
            updates.append(f"{field} = ${idx}")
            params.append(val)
            idx += 1
    if not updates: raise HTTPException(400, "No data")
    params.append(int(did))
    await db.execute(f"UPDATE users SET {', '.join(updates)} WHERE id = ${idx}", *params)
    row = await db.fetchrow("SELECT * FROM users WHERE id = $1", int(did))
    u = row_to_dict(row)
    u["id"] = str(u["id"])
    u.pop("password_hash", None)
    return u

@api_router.delete("/dealers/{did}")
async def delete_dealer(did: str, admin: dict = Depends(require_admin)):
    db = await get_pool()
    result = await db.execute("DELETE FROM users WHERE id = $1 AND role = 'dealer'", int(did))
    if result == "DELETE 0": raise HTTPException(404, "Not found")
    cache.invalidate("dealers", "chat", "stats")
    return {"message": "Deleted"}

# ─── WORKERS ───
@api_router.post("/workers")
async def create_worker(w: WorkerCreate, admin: dict = Depends(require_admin)):
    db = await get_pool()
    existing = await db.fetchrow("SELECT id FROM users WHERE email = $1", w.email.strip().lower())
    if existing: raise HTTPException(400, "Email mavjud")
    now = datetime.now(timezone.utc).isoformat()
    row = await db.fetchrow(
        "INSERT INTO users (name, email, password_hash, role, phone, address, credit_limit, debt, specialty, created_at) VALUES ($1,$2,$3,$4,$5,'',$6,0,$7,$8) RETURNING *",
        w.name, w.email.strip().lower(), hash_password(w.password), "worker", w.phone, 0, w.specialty, now
    )
    u = row_to_dict(row)
    u["id"] = str(u["id"])
    u.pop("password_hash", None)
    cache.invalidate("workers", "stats")
    return u

@api_router.get("/workers")
async def list_workers(admin: dict = Depends(require_admin)):
    cached = cache.get("workers_list")
    if cached: return cached
    db = await get_pool()
    rows = await db.fetch("SELECT * FROM users WHERE role = 'worker' ORDER BY created_at DESC")
    out = []
    for r in rows:
        u = row_to_dict(r)
        u["id"] = str(u["id"])
        u.pop("password_hash", None)
        out.append(u)
    cache.set("workers_list", out, 30)
    return out

@api_router.delete("/workers/{wid}")
async def delete_worker(wid: str, admin: dict = Depends(require_admin)):
    db = await get_pool()
    result = await db.execute("DELETE FROM users WHERE id = $1 AND role = 'worker'", int(wid))
    if result == "DELETE 0": raise HTTPException(404, "Not found")
    cache.invalidate("workers", "stats")
    return {"message": "Deleted"}

# ─── CATEGORIES ───
@api_router.post("/categories")
async def create_category(d: CategoryCreate, admin: dict = Depends(require_admin)):
    db = await get_pool()
    now = datetime.now(timezone.utc).isoformat()
    values = await filter_existing_fields(db, "categories", {
        "name": d.name,
        "description": d.description,
        "image_url": d.image_url,
        "created_at": now,
    })
    sql, params = build_insert_statement("categories", values)
    row = await db.fetchrow(sql, *params)
    c = row_to_dict(row); c["id"] = str(c["id"])
    cache.invalidate("categories", "materials")
    return c

@api_router.get("/categories")
async def list_categories(user: dict = Depends(get_current_user)):
    cached = cache.get("categories_list")
    if cached: return cached
    db = await get_pool()
    rows = await db.fetch("SELECT * FROM categories ORDER BY name ASC")
    out = []
    for r in rows:
        c = row_to_dict(r); c["id"] = str(c["id"])
        c["material_count"] = await db.fetchval(
            "SELECT COUNT(*) FROM materials WHERE category_id::text = $1",
            str(r["id"]),
        )
        out.append(c)
    cache.set("categories_list", out, 60)
    return out

@api_router.put("/categories/{cid}")
async def update_category(cid: str, d: CategoryUpdate, admin: dict = Depends(require_admin)):
    db = await get_pool()
    updates = []; params = []; idx = 1
    for field in ["name", "description", "image_url"]:
        val = getattr(d, field, None)
        if val is not None:
            updates.append(f"{field} = ${idx}"); params.append(val); idx += 1
    if not updates: raise HTTPException(400, "No data")
    params.append(str(cid))
    await db.execute(f"UPDATE categories SET {', '.join(updates)} WHERE id::text = ${idx}", *params)
    row = await db.fetchrow("SELECT * FROM categories WHERE id::text = $1", str(cid))
    c = row_to_dict(row); c["id"] = str(c["id"])
    cache.invalidate("categories", "materials")
    return c

@api_router.delete("/categories/{cid}")
async def delete_category(cid: str, admin: dict = Depends(require_admin)):
    db = await get_pool()
    mat_count = await db.fetchval(
        "SELECT COUNT(*) FROM materials WHERE category_id::text = $1",
        str(cid),
    )
    if mat_count > 0:
        raise HTTPException(400, f"Bu kategoriyada {mat_count} ta mahsulot bor. Avval mahsulotlarni ko'chiring.")
    result = await db.execute("DELETE FROM categories WHERE id::text = $1", str(cid))
    if result == "DELETE 0": raise HTTPException(404, "Not found")
    cache.invalidate("categories", "materials")
    return {"message": "Deleted"}

# ─── MATERIALS ───
@api_router.post("/materials")
async def create_material(d: MaterialCreate, admin: dict = Depends(require_admin)):
    db = await get_pool()
    now = datetime.now(timezone.utc).isoformat()
    values = await filter_existing_fields(db, "materials", {
        "name": d.name,
        "category": d.category,
        "category_id": d.category_id,
        "price_per_sqm": d.price_per_sqm,
        "stock_quantity": d.stock_quantity,
        "unit": d.unit,
        "description": d.description,
        "image_url": d.image_url,
        "created_at": now,
        "updated_at": now,
    })
    sql, params = build_insert_statement("materials", values)
    row = await db.fetchrow(sql, *params)
    m = row_to_dict(row); m["id"] = str(m["id"])
    if m.get("category_id"): m["category_id"] = str(m["category_id"])
    cache.invalidate("materials", "categories", "stats", "alerts")
    return m

@api_router.get("/materials")
async def list_materials(user: dict = Depends(get_current_user)):
    cached = cache.get("materials_list")
    if cached: return cached
    db = await get_pool()
    if await table_exists(db, "categories") and await column_exists(db, "materials", "category_id"):
        rows = await db.fetch(
            """
            SELECT m.*, c.name as category_name
            FROM materials m
            LEFT JOIN categories c ON m.category_id::text = c.id::text
            ORDER BY c.name ASC NULLS LAST, m.name ASC
            """
        )
    else:
        rows = await db.fetch("SELECT * FROM materials ORDER BY name ASC")
    out = []
    for r in rows:
        m = row_to_dict(r); m["id"] = str(m["id"])
        if m.get("category_id"): m["category_id"] = str(m["category_id"])
        m.setdefault("stock_quantity", 0)
        m.setdefault("unit", "kv.m")
        m.setdefault("category", "")
        m.setdefault("description", "")
        out.append(m)
    cache.set("materials_list", out, 60)
    return out

@api_router.get("/materials/by-category/{cid}")
async def list_materials_by_category(cid: str, user: dict = Depends(get_current_user)):
    cache_key = f"materials_cat_{cid}"
    cached = cache.get(cache_key)
    if cached: return cached
    db = await get_pool()
    rows = await db.fetch("SELECT * FROM materials WHERE category_id::text = $1 ORDER BY name ASC", str(cid))
    out = []
    for r in rows:
        m = row_to_dict(r); m["id"] = str(m["id"])
        if m.get("category_id"): m["category_id"] = str(m["category_id"])
        m.setdefault("stock_quantity", 0)
        m.setdefault("unit", "kv.m")
        out.append(m)
    cache.set(cache_key, out, 60)
    return out

@api_router.put("/materials/{mid}")
async def update_material(mid: str, d: MaterialUpdate, admin: dict = Depends(require_admin)):
    db = await get_pool()
    updates = []
    params = []
    idx = 1
    for field in ["name", "category", "price_per_sqm", "stock_quantity", "description", "image_url", "category_id"]:
        val = getattr(d, field, None)
        if val is not None and await column_exists(db, "materials", field):
            updates.append(f"{field} = ${idx}")
            params.append(val)
            idx += 1
    if not updates: raise HTTPException(400, "No data")
    params.append(str(mid))
    await db.execute(f"UPDATE materials SET {', '.join(updates)} WHERE id::text = ${idx}", *params)
    row = await db.fetchrow("SELECT * FROM materials WHERE id::text = $1", str(mid))
    m = row_to_dict(row)
    m["id"] = str(m["id"])
    cache.invalidate("materials", "categories", "alerts")
    return m

@api_router.delete("/materials/{mid}")
async def delete_material(mid: str, admin: dict = Depends(require_admin)):
    db = await get_pool()
    result = await db.execute("DELETE FROM materials WHERE id::text = $1", str(mid))
    if result == "DELETE 0": raise HTTPException(404, "Not found")
    cache.invalidate("materials", "categories", "stats", "alerts")
    return {"message": "Deleted"}

# ─── ORDERS ───
@api_router.post("/orders")
async def create_order(data: OrderCreate, user: dict = Depends(get_current_user)):
    if user.get("role") != "dealer":
        raise HTTPException(403, "Faqat dilerlar")
    if not data.items:
        raise HTTPException(400, "Buyurtmada kamida bitta mahsulot bo'lishi kerak")

    pool = await get_pool()
    items = []
    total_price = 0.0
    total_sqm = 0.0

    for it in data.items:
        order_item = build_order_item(it)
        if order_item["sqm"] <= 0:
            raise HTTPException(400, f"{it.material_name} uchun maydon 0 dan katta bo'lishi kerak")
        items.append(order_item)
        total_sqm += order_item["sqm"]
        total_price += order_item["price"]

    order_code = generate_order_code()
    now = datetime.now(timezone.utc).isoformat()

    async with pool.acquire() as conn:
        async with conn.transaction():
            status_value = await normalize_status_for_db(conn, "kutilmoqda")
            dealer_ids = await resolve_actor_ids(conn, user)
            dealer_ref = pick_reference_value(dealer_ids, await column_udt_name(conn, "orders", "dealer_id"))
            if await column_exists(conn, "orders", "dealer_id") and dealer_ref is None:
                raise HTTPException(400, "Diler profili topilmadi. Admin diler akkauntini qayta yaratib ko'ring.")

            if await column_exists(conn, "orders", "items"):
                order_values = await filter_existing_fields(conn, "orders", {
                    "order_code": order_code,
                    "dealer_id": dealer_ref,
                    "dealer_name": user.get("name", ""),
                    "items": json.dumps(items),
                    "total_sqm": round_money(total_sqm),
                    "total_price": round_money(total_price),
                    "status": status_value,
                    "notes": data.notes,
                    "rejection_reason": "",
                    "delivery_info": None,
                    "created_at": now,
                    "updated_at": now,
                })
                sql, params = build_insert_statement("orders", order_values)
                row = await conn.fetchrow(sql, *params)
            else:
                order_values = await filter_existing_fields(conn, "orders", {
                    "order_code": order_code,
                    "dealer_id": dealer_ref,
                    "dealer_name": user.get("name", ""),
                    "total_sqm": round_money(total_sqm),
                    "total_price": round_money(total_price),
                    "status": status_value,
                    "notes": data.notes,
                    "created_at": now,
                    "updated_at": now,
                })
                sql, params = build_insert_statement("orders", order_values)
                row = await conn.fetchrow(sql, *params)

                if await table_exists(conn, "order_items"):
                    material_udt = await column_udt_name(conn, "order_items", "material_id")
                    for item_index, item in enumerate(items):
                        item_values = await filter_existing_fields(conn, "order_items", {
                            "order_id": row["id"],
                            "material_id": pick_reference_value([item["material_id"]], material_udt),
                            "material_name": item["material_name"],
                            "width": item["width"],
                            "height": item["height"],
                            "quantity": item.get("quantity", 1),
                            "sqm": item["sqm"],
                            "unit_price": item["price_per_sqm"],
                            "total_price": item["price"],
                            "notes": item.get("notes", ""),
                            "worker_status": "pending",
                            "item_index": item_index,
                            "created_at": now,
                            "updated_at": now,
                        })
                        sql, params = build_insert_statement("order_items", item_values)
                        await conn.fetchrow(sql, *params)

            if await table_exists(conn, "users") and await column_exists(conn, "users", "debt"):
                await conn.execute(
                    "UPDATE users SET debt = COALESCE(debt, 0) + $1 WHERE id::text = $2",
                    round_money(total_price),
                    str(user["id"]),
                )

    cache.invalidate("orders", "stats", "reports")
    return await serialize_order(pool, row)

@api_router.get("/orders")
async def list_orders(user: dict = Depends(get_current_user)):
    cache_key = f"orders_{user['id']}_{user.get('role','')}"
    cached = cache.get(cache_key)
    if cached: return cached
    db = await get_pool()
    if user.get("role") == "dealer":
        actor_ids = await resolve_actor_ids(db, user)
        rows = await db.fetch(
            "SELECT * FROM orders WHERE dealer_id::text = ANY($1::text[]) ORDER BY created_at DESC",
            actor_ids,
        )
    elif user.get("role") == "worker" and await column_exists(db, "orders", "worker_id"):
        actor_ids = await resolve_actor_ids(db, user)
        rows = await db.fetch(
            "SELECT * FROM orders WHERE worker_id::text = ANY($1::text[]) ORDER BY created_at DESC",
            actor_ids,
        )
    else:
        rows = await db.fetch("SELECT * FROM orders ORDER BY created_at DESC")
    out = await serialize_orders(db, rows)
    cache.set(cache_key, out, 15)
    return out

@api_router.get("/orders/{oid}")
async def get_order(oid: str, user: dict = Depends(get_current_user)):
    db = await get_pool()
    o = await fetch_order_row(db, oid)
    if not o: raise HTTPException(404, "Not found")
    o = await serialize_order(db, o)
    if user.get("role") == "dealer":
        actor_ids = await resolve_actor_ids(db, user)
        if str(o["dealer_id"]) not in actor_ids:
            raise HTTPException(403)
    return o

@api_router.put("/orders/{oid}/status")
async def update_order_status(oid: str, data: OrderStatusUpdate, admin: dict = Depends(require_admin)):
    valid = ["kutilmoqda","tasdiqlangan","tayyorlanmoqda","tayyor","yetkazilmoqda","yetkazildi","rad_etilgan"]
    if data.status not in valid:
        raise HTTPException(400, "Invalid status")

    pool = await get_pool()
    now = datetime.now(timezone.utc).isoformat()

    async with pool.acquire() as conn:
        async with conn.transaction():
            order = await fetch_order_row(conn, oid)
            if not order:
                raise HTTPException(404, "Not found")

            if data.status in INVENTORY_SYNC_STATUSES:
                await deduct_inventory_for_order(conn, order, now)

            changes = {
                "status": await normalize_status_for_db(conn, data.status),
                "updated_at": now,
            }
            if data.status == "rad_etilgan" and data.rejection_reason:
                changes["rejection_reason"] = data.rejection_reason
            await update_order_fields(conn, oid, changes)
            updated = await fetch_order_row(conn, oid)

    cache.invalidate("orders", "stats", "reports", "materials", "alerts")
    return await serialize_order(pool, updated)

# ─── WORKER: Assign item to worker ───
@api_router.put("/orders/{oid}/items/{item_idx}/assign")
async def assign_item_to_worker(oid: str, item_idx: int, data: AssignItemReq, admin: dict = Depends(require_admin)):
    db = await get_pool()
    order = await fetch_order_row(db, oid)
    if not order: raise HTTPException(404, "Order not found")
    worker = await db.fetchrow("SELECT * FROM users WHERE id::text = $1 AND role = 'worker'", str(data.worker_id))
    if not worker: raise HTTPException(404, "Worker not found")
    if await column_exists(db, "orders", "items"):
        items = parse_json_field(order.get("items"), default=[]) or []
        if item_idx >= len(items): raise HTTPException(400, "Invalid item index")
        items[item_idx]["assigned_worker_id"] = str(data.worker_id)
        items[item_idx]["assigned_worker_name"] = worker["name"]
        items[item_idx]["worker_status"] = "assigned"
        await update_order_fields(db, oid, {"items": json.dumps(items)})
    else:
        item_rows = await db.fetch(
            """
            SELECT *
            FROM order_items
            WHERE order_id::text = $1
            ORDER BY COALESCE(item_index, 0) ASC, created_at ASC
            """,
            str(oid),
        )
        if item_idx >= len(item_rows):
            raise HTTPException(400, "Invalid item index")
        worker_data = row_to_dict(worker)
        worker_data["id"] = str(worker_data["id"])
        worker_ids = await resolve_actor_ids(db, worker_data)
        worker_ref = pick_reference_value(worker_ids, await column_udt_name(db, "orders", "worker_id"))
        if await column_exists(db, "orders", "worker_id") and worker_ref is None:
            raise HTTPException(400, "Ishchi profili topilmadi. Ishchini qayta yaratib ko'ring.")
        await update_order_fields(db, oid, {"worker_id": worker_ref, "updated_at": datetime.now(timezone.utc).isoformat()})
    o = await fetch_order_row(db, oid)
    cache.invalidate("orders")
    return await serialize_order(db, o)

# ─── WORKER: Get my assigned items ───
@api_router.get("/worker/tasks")
async def get_worker_tasks(user: dict = Depends(get_current_user)):
    if user.get("role") != "worker": raise HTTPException(403)
    db = await get_pool()
    actor_ids = await resolve_actor_ids(db, user)
    rows = await db.fetch(
        """
        SELECT *
        FROM orders
        WHERE status::text IN ('tasdiqlangan', 'tayyorlanmoqda', 'tayyor')
        ORDER BY created_at DESC
        """
    )
    orders = await serialize_orders(db, rows)
    tasks = []
    for o in orders:
        order_worker_id = str(o.get("worker_id") or "")
        for idx, item in enumerate(o.get("items") or []):
            assigned_worker_id = str(item.get("assigned_worker_id") or order_worker_id or "")
            if assigned_worker_id and assigned_worker_id in actor_ids:
                tasks.append({
                    "order_id": str(o["id"]),
                    "order_code": o.get("order_code", ""),
                    "dealer_name": o.get("dealer_name", ""),
                    "item_index": idx,
                    "material_name": item.get("material_name", ""),
                    "width": item.get("width", 0),
                    "height": item.get("height", 0),
                    "sqm": item.get("sqm", 0),
                    "notes": item.get("notes", ""),
                    "worker_status": item.get("worker_status", "assigned"),
                    "created_at": o.get("created_at"),
                })
    return tasks

# ─── WORKER: Mark item as completed ───
@api_router.put("/worker/tasks/{oid}/{item_idx}/complete")
async def complete_worker_task(oid: str, item_idx: int, user: dict = Depends(get_current_user)):
    if user.get("role") != "worker":
        raise HTTPException(403)
    pool = await get_pool()
    now = datetime.now(timezone.utc).isoformat()
    actor_ids = await resolve_actor_ids(pool, user)
    async with pool.acquire() as conn:
        async with conn.transaction():
            order = await fetch_order_row(conn, oid)
            if not order:
                raise HTTPException(404)

            if await column_exists(conn, "orders", "items"):
                items = parse_json_field(order["items"], default=[]) or []
                if item_idx >= len(items):
                    raise HTTPException(400)
                assigned_worker_id = str(items[item_idx].get("assigned_worker_id") or "")
                if assigned_worker_id and assigned_worker_id not in actor_ids:
                    raise HTTPException(403, "Not your task")
                if items[item_idx].get("worker_status") == "completed":
                    current = await serialize_order(conn, order)
                    current["items"] = items
                    return current
                items[item_idx]["worker_status"] = "completed"
                changes = {"items": json.dumps(items), "updated_at": now}

                all_done = all(it.get("worker_status") == "completed" for it in items if it.get("assigned_worker_id"))
                if all_done:
                    refreshed_order = dict(row_to_dict(order))
                    refreshed_order["items"] = json.dumps(items)
                    await deduct_inventory_for_order(conn, refreshed_order, now)
                    changes["status"] = await normalize_status_for_db(conn, "tayyor")
                await update_order_fields(conn, oid, changes)
            else:
                item_rows = await conn.fetch(
                    """
                    SELECT *
                    FROM order_items
                    WHERE order_id::text = $1
                    ORDER BY COALESCE(item_index, 0) ASC, created_at ASC
                    """,
                    str(oid),
                )
                if item_idx >= len(item_rows):
                    raise HTTPException(400)
                order_worker_id = str(order.get("worker_id") or "")
                if order_worker_id and order_worker_id not in actor_ids:
                    raise HTTPException(403, "Not your task")
                item_row = row_to_dict(item_rows[item_idx])
                if str(item_row.get("worker_status") or "") == "completed":
                    return await serialize_order(conn, order)

                item_updates = {"worker_status": "completed", "updated_at": now}
                updates = []
                params = []
                idx = 1
                for field, value in item_updates.items():
                    if await column_exists(conn, "order_items", field):
                        updates.append(f"{field} = ${idx}")
                        params.append(value)
                        idx += 1
                params.append(str(item_row["id"]))
                await conn.execute(
                    f"UPDATE order_items SET {', '.join(updates)} WHERE id::text = ${idx}",
                    *params,
                )

                remaining = await conn.fetchval(
                    """
                    SELECT COUNT(*)
                    FROM order_items
                    WHERE order_id::text = $1 AND worker_status::text != 'completed'
                    """,
                    str(oid),
                )
                if int(remaining or 0) == 0:
                    refreshed_order = await fetch_order_row(conn, oid)
                    await deduct_inventory_for_order(conn, refreshed_order, now)
                    await update_order_fields(conn, oid, {
                        "status": await normalize_status_for_db(conn, "tayyor"),
                        "updated_at": now,
                    })

            updated = await fetch_order_row(conn, oid)

    cache.invalidate("orders", "stats", "reports", "materials", "alerts")
    return await serialize_order(pool, updated)
# ─── DELIVERY: Assign delivery info directly to order ───
@api_router.put("/orders/{oid}/delivery")
async def assign_delivery(oid: str, data: DeliveryInfoReq, admin: dict = Depends(require_admin)):
    pool = await get_pool()
    d_info = json.dumps({"driver_name": data.driver_name, "driver_phone": data.driver_phone, "plate_number": data.plate_number})
    now = datetime.now(timezone.utc).isoformat()
    async with pool.acquire() as conn:
        async with conn.transaction():
            order = await fetch_order_row(conn, oid)
            if not order:
                raise HTTPException(404, "Buyurtma topilmadi")
            await deduct_inventory_for_order(conn, order, now)
            await update_order_fields(conn, oid, {
                "delivery_info": d_info,
                "status": await normalize_status_for_db(conn, "yetkazilmoqda"),
                "updated_at": now,
            })
            updated = await fetch_order_row(conn, oid)
    cache.invalidate("orders", "stats", "materials", "alerts")
    return await serialize_order(pool, updated)

# ─── DELIVERY: Admin confirms delivery ───
@api_router.put("/orders/{oid}/confirm-delivery")
async def confirm_delivery(oid: str, admin: dict = Depends(require_admin)):
    pool = await get_pool()
    now = datetime.now(timezone.utc).isoformat()
    async with pool.acquire() as conn:
        async with conn.transaction():
            order = await fetch_order_row(conn, oid)
            if not order:
                raise HTTPException(404, "Buyurtma topilmadi")
            await deduct_inventory_for_order(conn, order, now)
            await update_order_fields(conn, oid, {
                "status": await normalize_status_for_db(conn, "yetkazildi"),
                "updated_at": now,
            })
            updated = await fetch_order_row(conn, oid)
    cache.invalidate("orders", "stats", "reports", "materials", "alerts")
    return await serialize_order(pool, updated)

# ─── CHAT ───
@api_router.post("/messages")
async def send_message(data: MessageCreate, user: dict = Depends(get_current_user)):
    db = await get_pool()
    now = datetime.now(timezone.utc).isoformat()
    row = await db.fetchrow(
        "INSERT INTO messages (sender_id, sender_name, sender_role, receiver_id, text, read, created_at) VALUES ($1,$2,$3,$4,$5,$6,$7) RETURNING *",
        int(user["id"]), user.get("name", ""), user.get("role", ""), int(data.receiver_id), data.text, False, now
    )
    m = row_to_dict(row)
    m["id"] = str(m["id"])
    m["sender_id"] = str(m["sender_id"])
    m["receiver_id"] = str(m["receiver_id"])
    cache.invalidate("chat")
    return m

@api_router.get("/messages/{pid}")
async def get_messages(pid: str, user: dict = Depends(get_current_user)):
    db = await get_pool()
    rows = await db.fetch(
        "SELECT * FROM messages WHERE (sender_id = $1 AND receiver_id = $2) OR (sender_id = $2 AND receiver_id = $1) ORDER BY created_at ASC",
        int(user["id"]), int(pid)
    )
    out = []
    for r in rows:
        m = row_to_dict(r)
        m["id"] = str(m["id"])
        m["sender_id"] = str(m["sender_id"])
        m["receiver_id"] = str(m["receiver_id"])
        out.append(m)
    await db.execute("UPDATE messages SET read = TRUE WHERE sender_id = $1 AND receiver_id = $2 AND read = FALSE", int(pid), int(user["id"]))
    return out

@api_router.get("/chat/partners")
async def get_chat_partners(user: dict = Depends(get_current_user)):
    db = await get_pool()
    if user.get("role") == "admin":
        rows = await db.fetch("SELECT * FROM users WHERE role = 'dealer' ORDER BY name")
        out = []
        for r in rows:
            d = row_to_dict(r)
            d["id"] = str(d["id"])
            d.pop("password_hash", None)
            lm = await db.fetchrow(
                "SELECT text, created_at FROM messages WHERE (sender_id = $1 AND receiver_id = $2) OR (sender_id = $2 AND receiver_id = $1) ORDER BY created_at DESC LIMIT 1",
                int(user["id"]), r["id"]
            )
            uc = await db.fetchval(
                "SELECT COUNT(*) FROM messages WHERE sender_id = $1 AND receiver_id = $2 AND read = FALSE",
                r["id"], int(user["id"])
            )
            d["last_message"] = lm["text"] if lm else ""
            d["last_message_time"] = lm["created_at"] if lm else ""
            d["unread_count"] = uc or 0
            out.append(d)
        return out
    else:
        admin = await db.fetchrow("SELECT * FROM users WHERE role = 'admin' LIMIT 1")
        if not admin: return []
        a = row_to_dict(admin)
        a["id"] = str(a["id"])
        a.pop("password_hash", None)
        lm = await db.fetchrow(
            "SELECT text, created_at FROM messages WHERE (sender_id = $1 AND receiver_id = $2) OR (sender_id = $2 AND receiver_id = $1) ORDER BY created_at DESC LIMIT 1",
            int(user["id"]), admin["id"]
        )
        uc = await db.fetchval(
            "SELECT COUNT(*) FROM messages WHERE sender_id = $1 AND receiver_id = $2 AND read = FALSE",
            admin["id"], int(user["id"])
        )
        a["last_message"] = lm["text"] if lm else ""
        a["last_message_time"] = lm["created_at"] if lm else ""
        a["unread_count"] = uc or 0
        return [a]

# ─── IMAGE UPLOAD ───
@api_router.post("/upload-image")
async def upload_image(file: UploadFile = File(...), user: dict = Depends(require_admin)):
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(400, "Faqat rasm fayllari ruxsat etiladi")
    ext = file.filename.split(".")[-1] if file.filename and "." in file.filename else "jpg"
    filename = f"{uuid.uuid4().hex}.{ext}"
    filepath = UPLOAD_DIR / filename
    async with aiofiles.open(filepath, "wb") as f:
        content = await file.read()
        await f.write(content)
    image_url = f"/api/uploads/{filename}"
    return {"image_url": image_url}

# ─── STATISTICS ───
@api_router.get("/statistics")
async def get_statistics(admin: dict = Depends(require_admin)):
    cached = cache.get("stats_all")
    if cached: return cached
    db = await get_pool()
    people_table = await get_people_table(db)
    total_revenue = await db.fetchval(
        """
        SELECT COALESCE(SUM(total_price), 0)
        FROM orders
        WHERE status::text IN ('tasdiqlangan', 'tayyorlanmoqda', 'tayyor', 'yetkazilmoqda', 'yetkazildi')
        """
    )
    result = {
        "total_orders": await db.fetchval("SELECT COUNT(*) FROM orders"),
        "pending_orders": await db.fetchval("SELECT COUNT(*) FROM orders WHERE status::text = 'kutilmoqda'"),
        "approved_orders": await db.fetchval("SELECT COUNT(*) FROM orders WHERE status::text = 'tasdiqlangan'"),
        "preparing_orders": await db.fetchval("SELECT COUNT(*) FROM orders WHERE status::text = 'tayyorlanmoqda'"),
        "ready_orders": await db.fetchval("SELECT COUNT(*) FROM orders WHERE status::text = 'tayyor'"),
        "delivering_orders": await db.fetchval("SELECT COUNT(*) FROM orders WHERE status::text = 'yetkazilmoqda'"),
        "delivered_orders": await db.fetchval("SELECT COUNT(*) FROM orders WHERE status::text = 'yetkazildi'"),
        "rejected_orders": await db.fetchval("SELECT COUNT(*) FROM orders WHERE status::text = 'rad_etilgan'"),
        "total_dealers": await db.fetchval(f"SELECT COUNT(*) FROM {people_table} WHERE role::text = 'dealer'") if people_table else 0,
        "total_workers": await db.fetchval(f"SELECT COUNT(*) FROM {people_table} WHERE role::text = 'worker'") if people_table else 0,
        "total_materials": await db.fetchval("SELECT COUNT(*) FROM materials"),
        "total_revenue": round(float(total_revenue), 2),
    }
    cache.set("stats_all", result, 30)
    return result

# ─── SEED & STARTUP ───
async def create_tables(db):
    await db.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL DEFAULT '',
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'dealer',
            phone TEXT DEFAULT '',
            address TEXT DEFAULT '',
            credit_limit FLOAT DEFAULT 0,
            debt FLOAT DEFAULT 0,
            specialty TEXT DEFAULT '',
            created_at TEXT DEFAULT ''
        )
    """)
    await db.execute("""
        CREATE TABLE IF NOT EXISTS categories (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            description TEXT DEFAULT '',
            image_url TEXT DEFAULT '',
            created_at TEXT DEFAULT ''
        )
    """)
    await db.execute("""
        CREATE TABLE IF NOT EXISTS materials (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            category TEXT NOT NULL DEFAULT '',
            category_id INTEGER REFERENCES categories(id) ON DELETE SET NULL,
            price_per_sqm FLOAT NOT NULL DEFAULT 0,
            stock_quantity FLOAT NOT NULL DEFAULT 0,
            unit TEXT DEFAULT 'kv.m',
            description TEXT DEFAULT '',
            image_url TEXT DEFAULT '',
            created_at TEXT DEFAULT ''
        )
    """)
    # Add category_id column if not exists (migration)
    try:
        await db.execute("ALTER TABLE materials ADD COLUMN IF NOT EXISTS category_id INTEGER REFERENCES categories(id) ON DELETE SET NULL")
    except Exception:
        pass
    await db.execute("""
        CREATE TABLE IF NOT EXISTS orders (
            id SERIAL PRIMARY KEY,
            order_code TEXT DEFAULT '',
            dealer_id INTEGER REFERENCES users(id),
            dealer_name TEXT DEFAULT '',
            items TEXT DEFAULT '[]',
            total_sqm FLOAT DEFAULT 0,
            total_price FLOAT DEFAULT 0,
            status TEXT DEFAULT 'kutilmoqda',
            inventory_deducted BOOLEAN NOT NULL DEFAULT FALSE,
            notes TEXT DEFAULT '',
            rejection_reason TEXT DEFAULT '',
            delivery_info TEXT,
            created_at TEXT DEFAULT '',
            updated_at TEXT DEFAULT ''
        )
    """)
    try:
        await db.execute("ALTER TABLE orders ADD COLUMN IF NOT EXISTS inventory_deducted BOOLEAN NOT NULL DEFAULT FALSE")
    except Exception:
        pass
    await db.execute("""
        CREATE TABLE IF NOT EXISTS messages (
            id SERIAL PRIMARY KEY,
            sender_id INTEGER REFERENCES users(id),
            sender_name TEXT DEFAULT '',
            sender_role TEXT DEFAULT '',
            receiver_id INTEGER REFERENCES users(id),
            text TEXT DEFAULT '',
            read BOOLEAN DEFAULT FALSE,
            created_at TEXT DEFAULT ''
        )
    """)
    await db.execute("""
        CREATE TABLE IF NOT EXISTS payments (
            id SERIAL PRIMARY KEY,
            dealer_id INTEGER REFERENCES users(id),
            amount FLOAT NOT NULL DEFAULT 0,
            note TEXT DEFAULT '',
            created_at TEXT DEFAULT ''
        )
    """)
    await db.execute("""
        CREATE TABLE IF NOT EXISTS user_settings (
            id SERIAL PRIMARY KEY,
            user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            theme TEXT NOT NULL DEFAULT 'system',
            notifications BOOLEAN NOT NULL DEFAULT TRUE,
            created_at TEXT DEFAULT '',
            updated_at TEXT DEFAULT ''
        )
    """)

async def seed_admin(db):
    email = os.environ.get("ADMIN_EMAIL", "admin@curtain.uz")
    pw = os.environ.get("ADMIN_PASSWORD", "admin123")
    now = datetime.now(timezone.utc).isoformat()
    ex = await db.fetchrow("SELECT * FROM users WHERE email = $1", email)
    if not ex:
        await db.execute(
            "INSERT INTO users (name, email, password_hash, role, phone, address, credit_limit, debt, specialty, created_at) VALUES ($1,$2,$3,$4,'','',0,0,'',$5)",
            "Admin", email, hash_password(pw), "admin", now
        )
        logger.info(f"Admin yaratildi: {email}")
    elif not verify_password(pw, ex["password_hash"]):
        await db.execute("UPDATE users SET password_hash = $1 WHERE email = $2", hash_password(pw), email)

    mat_count = await db.fetchval("SELECT COUNT(*) FROM materials")
    if mat_count == 0:
        # Create default categories
        cat_count = await db.fetchval("SELECT COUNT(*) FROM categories")
        if cat_count == 0:
            cats = [
                ("Parda", "Har xil parda turlari"),
                ("Jalyuzi", "Gorizontal va vertikal jalyuzilar"),
                ("Aksessuar", "Karniz, gardina va boshqa aksessuarlar"),
            ]
            for c in cats:
                await db.execute("INSERT INTO categories (name, description, created_at) VALUES ($1,$2,$3)", c[0], c[1], now)
            logger.info("Kategoriyalar yaratildi")

        parda_id = await db.fetchval("SELECT id FROM categories WHERE name = 'Parda'")
        jalyuzi_id = await db.fetchval("SELECT id FROM categories WHERE name = 'Jalyuzi'")

        materials_data = [
            ("Blackout Parda", "Parda", parda_id, 7.0, 500, "kv.m", "Yorug'lik o'tkazmaydigan parda", "https://images.pexels.com/photos/4814070/pexels-photo-4814070.jpeg?auto=compress&cs=tinysrgb&dpr=2&h=650&w=940"),
            ("Tull Parda", "Parda", parda_id, 3.5, 800, "kv.m", "Shaffof tull parda", "https://images.unsplash.com/photo-1574197635162-68e4b468e4e9?w=600"),
            ("Roller Jalyuzi", "Jalyuzi", jalyuzi_id, 10.0, 300, "kv.m", "Zamonaviy roller jalyuzi", "https://images.pexels.com/photos/19166538/pexels-photo-19166538.jpeg?auto=compress&cs=tinysrgb&dpr=2&h=650&w=940"),
            ("Gorizontal Jalyuzi", "Jalyuzi", jalyuzi_id, 8.0, 400, "kv.m", "Alyuminiy gorizontal jalyuzi", "https://images.unsplash.com/photo-1603299938527-d035bc6fc2c8?w=600"),
            ("Vertikal Jalyuzi", "Jalyuzi", jalyuzi_id, 6.0, 350, "kv.m", "Ofis uchun vertikal jalyuzi", "https://images.pexels.com/photos/8955198/pexels-photo-8955198.jpeg?auto=compress&cs=tinysrgb&dpr=2&h=650&w=940"),
            ("Rimskaya Parda", "Parda", parda_id, 9.0, 200, "kv.m", "Premium rimskaya parda", "https://images.unsplash.com/photo-1729277980958-092c5e9e2ea4?w=600"),
        ]
        for m in materials_data:
            await db.execute(
                "INSERT INTO materials (name, category, category_id, price_per_sqm, stock_quantity, unit, description, image_url, created_at) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9)",
                m[0], m[1], m[2], m[3], m[4], m[5], m[6], m[7], now
            )
        logger.info("Materiallar yaratildi")

    if not await db.fetchrow("SELECT id FROM users WHERE email = 'dealer@test.uz'"):
        await db.execute(
            "INSERT INTO users (name, email, password_hash, role, phone, address, credit_limit, debt, specialty, created_at) VALUES ($1,$2,$3,$4,$5,$6,$7,0,'',$8)",
            "Test Diler", "dealer@test.uz", hash_password("dealer123"), "dealer", "+998901234567", "Toshkent, Yunusobod", 5000.0, now
        )
        logger.info("Demo diler yaratildi")

    if not await db.fetchrow("SELECT id FROM users WHERE email = 'worker@test.uz'"):
        await db.execute(
            "INSERT INTO users (name, email, password_hash, role, phone, address, credit_limit, debt, specialty, created_at) VALUES ($1,$2,$3,$4,$5,'',$6,0,$7,$8)",
            "Aziz Ishchi", "worker@test.uz", hash_password("worker123"), "worker", "+998901112233", 0.0, "Jalyuzi o'rnatish", now
        )
        logger.info("Demo ishchi yaratildi")

# ─── REPORTS - Hisobot tizimi ───
@api_router.get("/reports")
async def get_reports(admin: dict = Depends(require_admin)):
    cached = cache.get("reports_all")
    if cached: return cached
    db = await get_pool()
    now = datetime.now(timezone.utc)

    # Weekly & Monthly revenue
    week_ago = (now - timedelta(days=7)).isoformat()
    month_ago = (now - timedelta(days=30)).isoformat()

    weekly_revenue = await db.fetchval(
        "SELECT COALESCE(SUM(total_price), 0) FROM orders WHERE created_at >= $1 AND status::text NOT IN ('rad_etilgan')", week_ago
    )
    monthly_revenue = await db.fetchval(
        "SELECT COALESCE(SUM(total_price), 0) FROM orders WHERE created_at >= $1 AND status::text NOT IN ('rad_etilgan')", month_ago
    )
    total_revenue = await db.fetchval(
        "SELECT COALESCE(SUM(total_price), 0) FROM orders WHERE status::text NOT IN ('rad_etilgan')"
    )
    weekly_orders = await db.fetchval("SELECT COUNT(*) FROM orders WHERE created_at >= $1", week_ago)
    monthly_orders = await db.fetchval("SELECT COUNT(*) FROM orders WHERE created_at >= $1", month_ago)
    total_orders = await db.fetchval("SELECT COUNT(*) FROM orders")

    # Top selling materials (from order items)
    mat_stats: dict = {}
    if await column_exists(db, "orders", "items"):
        all_orders = await db.fetch("SELECT items FROM orders WHERE status::text NOT IN ('rad_etilgan')")
        for row in all_orders:
            items = parse_json_field(row["items"], default=[]) or []
            for it in items:
                name = it.get("material_name", "Noma'lum")
                sqm = float(it.get("sqm", 0) or 0)
                price = float(it.get("price", it.get("total_price", 0)) or 0)
                if name not in mat_stats:
                    mat_stats[name] = {"name": name, "total_sqm": 0, "total_price": 0, "count": 0}
                mat_stats[name]["total_sqm"] += sqm
                mat_stats[name]["total_price"] += price
                mat_stats[name]["count"] += 1
    elif await table_exists(db, "order_items"):
        item_rows = await db.fetch(
            """
            SELECT oi.material_name, oi.sqm, oi.total_price
            FROM order_items oi
            JOIN orders o ON oi.order_id::text = o.id::text
            WHERE o.status::text NOT IN ('rad_etilgan')
            """
        )
        for row in item_rows:
            name = row.get("material_name") or "Noma'lum"
            sqm = float(row.get("sqm") or 0)
            price = float(row.get("total_price") or 0)
            if name not in mat_stats:
                mat_stats[name] = {"name": name, "total_sqm": 0, "total_price": 0, "count": 0}
            mat_stats[name]["total_sqm"] += sqm
            mat_stats[name]["total_price"] += price
            mat_stats[name]["count"] += 1

    top_materials = sorted(mat_stats.values(), key=lambda x: x["total_price"], reverse=True)[:5]

    # Top dealers
    dealer_rows = await db.fetch(
        """
        SELECT dealer_id::text AS dealer_id, COUNT(id) AS order_count, COALESCE(SUM(total_price), 0) AS revenue
        FROM orders
        WHERE status::text NOT IN ('rad_etilgan')
        GROUP BY dealer_id::text
        ORDER BY revenue DESC
        LIMIT 5
        """
    )
    dealer_name_map = await load_user_name_map(db, [str(row["dealer_id"]) for row in dealer_rows])
    top_dealers = [
        {
            "name": dealer_name_map.get(str(row["dealer_id"]), "Diler"),
            "orders": int(row["order_count"] or 0),
            "revenue": round(float(row["revenue"] or 0), 2),
        }
        for row in dealer_rows
    ]

    # Daily orders for last 7 days
    daily = []
    for i in range(6, -1, -1):
        day_start = (now - timedelta(days=i)).replace(hour=0, minute=0, second=0).isoformat()
        day_end = (now - timedelta(days=i)).replace(hour=23, minute=59, second=59).isoformat()
        cnt = await db.fetchval("SELECT COUNT(*) FROM orders WHERE created_at >= $1 AND created_at <= $2", day_start, day_end)
        rev = await db.fetchval("SELECT COALESCE(SUM(total_price), 0) FROM orders WHERE created_at >= $1 AND created_at <= $2 AND status::text NOT IN ('rad_etilgan')", day_start, day_end)
        day_label = (now - timedelta(days=i)).strftime("%d.%m")
        daily.append({"day": day_label, "orders": cnt, "revenue": round(float(rev), 2)})

    result = {
        "weekly_revenue": round(float(weekly_revenue), 2),
        "monthly_revenue": round(float(monthly_revenue), 2),
        "total_revenue": round(float(total_revenue), 2),
        "weekly_orders": weekly_orders,
        "monthly_orders": monthly_orders,
        "total_orders": total_orders,
        "top_materials": top_materials,
        "top_dealers": top_dealers,
        "daily": daily,
    }
    cache.set("reports_all", result, 60)
    return result

# ─── LOW STOCK ALERTS ───
@api_router.get("/alerts/low-stock")
async def get_low_stock(admin: dict = Depends(require_admin)):
    cached = cache.get("alerts_low_stock")
    if cached: return cached
    db = await get_pool()
    if not await column_exists(db, "materials", "stock_quantity"):
        cache.set("alerts_low_stock", [], 60)
        return []
    rows = await db.fetch("SELECT * FROM materials WHERE stock_quantity < 10 ORDER BY stock_quantity ASC")
    out = []
    for r in rows:
        m = row_to_dict(r)
        m["id"] = str(m["id"])
        out.append(m)
    cache.set("alerts_low_stock", out, 60)
    return out

# ─── EXCEL EXPORT ───
@api_router.get("/reports/export-orders")
async def export_orders_excel(admin: dict = Depends(require_admin)):
    db = await get_pool()
    order_rows = await db.fetch("SELECT * FROM orders ORDER BY created_at DESC")
    orders = await serialize_orders(db, order_rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Buyurtmalar"

    # Styling
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='FF453A', end_color='FF453A', fill_type='solid')
    border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD'),
    )

    headers = ['#', 'Buyurtma kodi', 'Diler', 'Mahsulotlar', 'Jami kv.m', 'Jami narx ($)', 'Status', 'Sana']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for idx, order in enumerate(orders, 1):
        items = order.get("items") or []
        item_names = ", ".join([
            f'{it.get("material_name", "")} ({it.get("width", 0)}x{it.get("height", 0)}m, {it.get("sqm", 0)} kv.m)'
            for it in items
        ])
        status_map = {"kutilmoqda": "Kutilmoqda", "tasdiqlangan": "Tasdiqlangan", "tayyorlanmoqda": "Tayyorlanmoqda", "tayyor": "Tayyor", "yetkazilmoqda": "Yetkazilmoqda", "yetkazildi": "Yetkazildi", "rad_etilgan": "Rad etilgan"}

        row = [
            idx,
            order["order_code"],
            order.get("dealer_name", ""),
            item_names,
            round_money(order.get("total_sqm", 0)),
            round_money(order.get("total_price", 0)),
            status_map.get(order.get("status", ""), order.get("status", "")),
            format_export_datetime(order.get("created_at")),
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=idx+1, column=col, value=val)
            cell.border = border
            if col in [5, 6]:
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    today = datetime.now(timezone.utc).strftime('%Y%m%d')

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=buyurtmalar_{today}.xlsx"}
    )

# ─── DEALER PAYMENTS (To'lovlar) ───
@api_router.post("/dealers/{did}/payment")
async def add_dealer_payment(did: str, data: PaymentCreate, admin: dict = Depends(require_admin)):
    if data.amount <= 0: raise HTTPException(400, "Summa 0 dan katta bo'lishi kerak")
    db = await get_pool()
    dealer = await db.fetchrow("SELECT * FROM users WHERE id = $1 AND role = 'dealer'", int(did))
    if not dealer: raise HTTPException(404, "Diler topilmadi")
    now = datetime.now(timezone.utc).isoformat()
    await db.execute(
        "INSERT INTO payments (dealer_id, amount, note, created_at) VALUES ($1,$2,$3,$4)",
        int(did), data.amount, data.note, now
    )
    new_debt = max(0, (dealer["debt"] or 0) - data.amount)
    await db.execute("UPDATE users SET debt = $1 WHERE id = $2", new_debt, int(did))
    cache.invalidate("dealers", "stats")
    return {"message": "To'lov qabul qilindi", "new_debt": round(new_debt, 2), "paid": data.amount}

@api_router.get("/dealers/{did}/payments")
async def get_dealer_payments(did: str, admin: dict = Depends(require_admin)):
    db = await get_pool()
    rows = await db.fetch("SELECT * FROM payments WHERE dealer_id = $1 ORDER BY created_at DESC", int(did))
    out = []
    for r in rows:
        p = row_to_dict(r)
        p["id"] = str(p["id"])
        p["dealer_id"] = str(p["dealer_id"])
        out.append(p)
    return out

# ─── HEALTH CHECK ───
@api_router.get("/health")
async def health_check():
    started_at = _time.perf_counter()
    try:
        db = await get_pool()
        await db.fetchval("SELECT 1")
        latency_ms = round((_time.perf_counter() - started_at) * 1000, 2)
        return {"status": "ok", "database": "connected", "latency_ms": latency_ms, "time": datetime.now(timezone.utc).isoformat()}
    except Exception as e:
        latency_ms = round((_time.perf_counter() - started_at) * 1000, 2)
        return JSONResponse(
            status_code=503,
            content={"status": "error", "database": str(e), "latency_ms": latency_ms},
        )

@api_router.get("/settings/me")
async def get_settings(user: dict = Depends(get_current_user)):
    db = await get_pool()
    row = await db.fetchrow(
        "SELECT theme, notifications, created_at, updated_at FROM user_settings WHERE user_id = $1",
        int(user["id"]),
    )
    if not row:
        return {"theme": "system", "notifications": True}
    return dict(row)

@api_router.put("/settings/me")
async def update_settings(request: Request, user: dict = Depends(get_current_user)):
    payload = await request.json()
    theme = str(payload.get("theme", "system")).strip().lower()
    notifications = bool(payload.get("notifications", True))
    if theme not in {"system", "light", "dark"}:
        raise HTTPException(400, "Invalid theme")

    db = await get_pool()
    now = datetime.now(timezone.utc).isoformat()
    row = await db.fetchrow(
        """
        INSERT INTO user_settings (user_id, theme, notifications, created_at, updated_at)
        VALUES ($1, $2, $3, $4, $4)
        ON CONFLICT (user_id)
        DO UPDATE SET theme = EXCLUDED.theme, notifications = EXCLUDED.notifications, updated_at = EXCLUDED.updated_at
        RETURNING theme, notifications, created_at, updated_at
        """,
        int(user["id"]),
        theme,
        notifications,
        now,
    )
    return dict(row)

# ─── KEEP ALIVE - PostgreSQL uxlab qolmasligi uchun ───
async def keep_alive_task():
    """Backend va DB uyg'oq turishi uchun qisqa intervalda ping yuboradi."""
    while True:
        try:
            await asyncio.sleep(KEEP_ALIVE_INTERVAL_SECONDS)
            db = await get_pool()
            await db.fetchval("SELECT 1")
            logger.debug("Keep-alive: PostgreSQL OK")
        except asyncio.CancelledError:
            break
        except Exception as e:
            logger.warning("Keep-alive ping xatolik: %s", e)

@app.on_event("startup")
async def startup():
    global pool
    pool = await asyncpg.create_pool(DATABASE_URL, **asyncpg_pool_kwargs())
    async with pool.acquire() as conn:
        await create_tables(conn)
        await create_indexes(conn)
        await seed_admin(conn)
    asyncio.create_task(keep_alive_task())
    logger.info("Muvaffaqiyat: API tayyor, PostgreSQL ulandi, keep-alive yoqildi (bu xato emas).")

@app.on_event("shutdown")
async def shutdown():
    global pool
    if pool:
        await pool.close()

app.include_router(api_router)
app.add_middleware(GZipMiddleware, minimum_size=800)
app.add_middleware(CORSMiddleware, allow_credentials=True, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
